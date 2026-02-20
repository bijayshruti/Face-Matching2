import streamlit as st
import cv2
import numpy as np
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
import plotly.graph_objects as go
import tempfile
import os
import shutil

# Page configuration
st.set_page_config(
    page_title="Face Grouping Pro",
    page_icon="👥",
    layout="wide"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #4b8bec 0%, #2a5f9e 100%);
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        color: white;
        text-align: center;
    }
    .info-box {
        background-color: #f0f8ff;
        padding: 1rem;
        border-radius: 5px;
        border-left: 5px solid #4b8bec;
        margin-bottom: 1rem;
    }
    .success-box {
        background-color: #dff0d8;
        padding: 1rem;
        border-radius: 5px;
        border-left: 5px solid #5cb85c;
        margin-bottom: 1rem;
    }
    .error-box {
        background-color: #f2dede;
        padding: 1rem;
        border-radius: 5px;
        border-left: 5px solid #d9534f;
        margin-bottom: 1rem;
    }
</style>
""", unsafe_allow_html=True)

# Header
st.markdown("""
<div class="main-header">
    <h1>👥 Face Grouping Professional</h1>
    <p>Upload images to group similar faces</p>
</div>
""", unsafe_allow_html=True)

# Check for face_recognition
try:
    import face_recognition
    FACE_RECOGNITION_AVAILABLE = True
except ImportError:
    FACE_RECOGNITION_AVAILABLE = False
    st.markdown("""
    <div class="error-box">
        <h4>⚠️ Face Recognition Library Not Available</h4>
        <p>The face recognition library could not be loaded. Please check:</p>
        <ul>
            <li>Python version should be 3.9 or lower</li>
            <li>System dependencies are installed</li>
            <li>The app is using the correct Python version</li>
        </ul>
    </div>
    """, unsafe_allow_html=True)

# Initialize session state
if 'processed' not in st.session_state:
    st.session_state.processed = False

# File uploader
uploaded_files = st.file_uploader(
    "Choose images...",
    type=['jpg', 'jpeg', 'png'],
    accept_multiple_files=True
)

if uploaded_files:
    st.success(f"✅ {len(uploaded_files)} images uploaded!")
    
    # Preview
    st.markdown("### 👁️ Preview")
    cols = st.columns(4)
    for idx, file in enumerate(uploaded_files[:4]):
        with cols[idx]:
            image = Image.open(file)
            st.image(image, caption=file.name, use_column_width=True)

# Settings
threshold = st.slider(
    "Similarity Threshold",
    min_value=0.1,
    max_value=0.8,
    value=0.3,
    step=0.05
)

# Process button
if st.button("🚀 Start Processing", type="primary"):
    if not uploaded_files:
        st.error("Please upload images first!")
    elif not FACE_RECOGNITION_AVAILABLE:
        st.error("Face recognition is not available. Please check the error message above.")
    else:
        # Create a simple demo output for now
        st.markdown("""
        <div class="success-box">
            <h4>✅ Demo Mode</h4>
            <p>The app is working! To fix the face recognition:</p>
            <ol>
                <li>Make sure runtime.txt has <code>python-3.9.18</code></li>
                <li>Add packages.txt with build dependencies</li>
                <li>Restart the app on Streamlit Cloud</li>
            </ol>
        </div>
        """, unsafe_allow_html=True)
        
        # Create a sample Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Grouped Faces"
        
        ws['A1'] = "Filename"
        ws['B1'] = "Group"
        
        for idx, file in enumerate(uploaded_files[:5]):
            ws[f'A{idx+2}'] = file.name
            ws[f'B{idx+2}'] = f"Group {idx//2 + 1}"
        
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        st.download_button(
            label="📥 Download Sample Report",
            data=excel_bytes,
            file_name="sample_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# System info (for debugging)
with st.expander("🔧 System Information"):
    import sys
    st.write(f"Python version: {sys.version}")
    st.write(f"Face Recognition available: {FACE_RECOGNITION_AVAILABLE}")
    
    if not FACE_RECOGNITION_AVAILABLE:
        st.write("To fix this, ensure your repository has:")
        st.code("""
# runtime.txt
python-3.9.18

# packages.txt
build-essential
cmake
python3.9-dev

# requirements.txt (simplified)
streamlit==1.28.0
opencv-python-headless==4.8.1.78
numpy==1.24.3
Pillow==10.0.0
plotly==5.17.0
pandas==2.0.3
openpyxl==3.1.2
face-recognition==1.3.0
        """)

# Footer
st.markdown("---")
st.markdown("Designed by Bijay Paswan | v1.0")